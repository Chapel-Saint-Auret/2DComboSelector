"""Headless regression tests for export flows."""

from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QApplication

from combo_selector.ui.pages.export_page import ExportPage
from tests.helpers import build_ranked_model, get_fixture_path


def _set_checked_items(tree_widget, expected_items: list[str]) -> None:
    """Update a CheckableTreeList with the requested checked child labels."""
    expected = set(expected_items)
    tree_widget.tree.blockSignals(True)
    for child in tree_widget.children:
        child.setCheckState(0, Qt.Checked if child.text(0) in expected else Qt.Unchecked)
    tree_widget.tree.blockSignals(False)


class ExportPageTests(unittest.TestCase):
    """Cover table and figure export without a visible desktop session."""

    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.model = build_ranked_model(
            get_fixture_path("release_format_ranking.xlsx"),
            peak_capacity_sheet="1D peak capacity table",
            elution_sheet="Elution-Composition Range Table",
        )
        self.page = ExportPage(self.model)
        self.page.init_page([])

    def tearDown(self) -> None:
        self.page.deleteLater()
        self.temp_dir.cleanup()

    def test_export_tables_writes_selected_sheets(self) -> None:
        self.page.table_export_directory_lineEdit.setText(self.temp_dir.name)
        self.page.export_filename.setText("results_export")
        _set_checked_items(
            self.page.table_selection,
            ["2D Combination Table", "Overall Results Table"],
        )

        with patch(
            "combo_selector.ui.pages.export_page.QMessageBox.warning",
            side_effect=AssertionError("Unexpected warning dialog"),
        ):
            self.page.export_tables()

        output = Path(self.temp_dir.name) / "results_export.xlsx"
        self.assertTrue(output.exists())
        workbook = load_workbook(output)
        self.assertEqual(
            workbook.sheetnames,
            ["2D Combination Table", "Overall Results Table"],
        )

    def test_save_figure_list_creates_expected_png(self) -> None:
        self.page.figure_export_directory_lineEdit.setText(self.temp_dir.name)
        self.page.figure_folder_name_lineEdit.setText("ReleaseFigures")
        _set_checked_items(self.page.figure_type_chklist, ["Scatter"])
        _set_checked_items(self.page.figure_list_chklist, ["Set 1"])

        with patch(
            "combo_selector.ui.pages.export_page.QMessageBox.warning",
            side_effect=AssertionError("Unexpected warning dialog"),
        ):
            self.page.save_figure_list()

        output = Path(self.temp_dir.name) / "ReleaseFigures" / "Scatter" / "Set 1.png"
        self.assertTrue(output.exists())
        self.assertGreater(output.stat().st_size, 0)
